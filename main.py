import pyodbc
import os, shutil
from openpyxl import Workbook, load_workbook
import tkinter as tk
import tkinter.font as font
import tkinter.ttk as ttk
from tkinter import messagebox, StringVar, OptionMenu, ttk
from tkinter.filedialog import asksaveasfile 
import datetime, sv_ttk
import pywinstyles, sys

#Notes
"""

add todays date - done
Final account decom works -> M5,6,7
Symology -> M8
Access -> M9, M10, M11

TM/PM notes-> VMO2 report comments G8
"""
def apply_theme_to_titlebar(root):
    version = sys.getwindowsversion()

    if version.major == 10 and version.build >= 22000:
        # Set the title bar color to the background color on Windows 11 for better appearance
        pywinstyles.change_header_color(root, "#1c1c1c" if sv_ttk.get_theme() == "dark" else "#fafafa")
    elif version.major == 10:
        pywinstyles.apply_style(root, "dark" if sv_ttk.get_theme() == "dark" else "normal")

        # A hacky way to update the title bar's color on Windows 10 (it doesn't update instantly like on Windows 11)
        root.wm_attributes("-alpha", 0.99)
        root.wm_attributes("-alpha", 1)

def validate_input():
    user_input = combo.get()
    try:
        if int(user_input) in TEFCSR:
            createQuote(int(user_input))
        else:
            messagebox.showerror("Site Number Not Found!", "Please check that you have entered the correct site number.")
    except Exception as e:
        messagebox.showerror("Site Number Not Found!", "Please check that you have selected a site number.")
        print(e)

def createQuote(site_num, msgbox=True):
    i = TEFCSR.index(site_num)
    access_type = {
        "" : 0,
        "N/A": 0,
        "None": 0,
        "Ladder" : 0,
        "Ladders": 0,
        "Podium Steps" : 9,
        "Mobile Scaffolding" : 10,
        "Mobile Scaffold" : 10,
        "X Tower/Scaffold" : 10,
        "Xtower" : 10,
        "Scaffold": 10,
        "MEWP": 11,
        "Cherry Picker" : 11
    }
    #Set a default filename
    filename = f"NET - VMO2 - {str(site_num)} - {site_name[i]} - CS_V1.xlsx"

    #Prompt user with file save dialog
    filesave = asksaveasfile(filetypes = [('Excel Document', '*.xlsx')], defaultextension = [('Excel Document', '*.xlsx')], initialfile=filename)

    wb = load_workbook("bins/VMO2 CS TEMPLATE.xlsx")
    ws = wb["Summary"]
    ws.cell(4, 2).value = site_num
    ws.cell(5, 2).value = site_name[i]
    ws.cell(6, 2).value = site_address[i]
    ws.cell(7, 2).value = site_postcode[i]
    ws.cell(9, 2).value = datetime.datetime.today()
    ws.cell(22, 1).value = document_name[i]
    ws = wb["Ratecard"]
    if cost_recieved[i] == "N/A":
        ws.cell(8, 3).value = 0
    else:
        ws.cell(8, 3).value = cost_recieved[i]
        
    if site_access[i] != None:
        if access_type[site_access[i]] > 0:
            ws.cell(access_type[site_access[i]], 5).value = 1
    if split_decom[i] == True:
        ws.cell(6, 5).value = 1
    if ooh_decom[i] == True:
        ws.cell(7, 5).value = 1

    ws.cell(8, 7).value = notes[i]

    #Save workbook with name selected by user.
    wb.save(filesave.name)
    
    if msgbox:
        messagebox.showinfo("Quote Created!", "Quote for site: " + str(site_num) + " - " + site_name[i] + " has been created!")

def createAllQuotes():
    for id in TEFCSR:
        try:
            createQuote(id, False)
        except Exception as e:
            messagebox.showerror("Error", f"Site Skipped: {id}, {e}")

    messagebox.showinfo("Quotes Created", "Quotes have been created for every site.")

server_loc = "T:\\New Server Structure\\Finance\\VMO2 Decoms SP.accdb"
db_loc = os.getcwd() + "\\VMO2 Decoms SP.accdb"

print("Downloading Database...")
shutil.copyfile(server_loc, "VMO2 Decoms SP.accdb")

conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + db_loc +  r';')
cursor = conn.cursor()
cursor.execute('SELECT TEFCSR, [Site Name], [Site Address], Postcode, [Survey Report Name], [TM Quote], [Access Equipment Required], [Split Decom], [OOH Decom], [TM / PM notes] FROM TblVMO2DecomMAIN')
dbdata = cursor.fetchall()

TEFCSR = []
for row in dbdata:
    try:
        TEFCSR.append(int(row[0]))
    except:
        continue

site_name = [row[1] for row in dbdata]
site_address = [row[2] for row in dbdata]
site_postcode = [row[3] for row in dbdata]
document_name = [row[4] for row in dbdata]
cost_recieved = [row[5] for row in dbdata]
site_access = [row[6] for row in dbdata]
split_decom = [row[7] for row in dbdata]
ooh_decom =  [row[8] for row in dbdata]
notes = [row[9] for row in dbdata]

root = tk.Tk()
root.iconbitmap("imgs/NETCS_Logo_Square_White.ico")
root.geometry("300x130")
root.resizable(False, False)
root.title("VMO2 Quote Generator")

label = tk.Label(root, text="Select Site Number", font=font.Font(size=12))
label.pack(pady=10)

combo = ttk.Combobox(
    state="readonly",
    values=sorted(TEFCSR)
)
combo.pack()
button = ttk.Button(root, text="Submit", command=validate_input)
button.pack(pady=10)

#button2 = tk.Button(root, text="Create Quote for all sites", command=createAllQuotes, font=font.Font(size=12))
#button2.pack(pady=10)

sv_ttk.set_theme("dark")
apply_theme_to_titlebar(root)

root.mainloop()